
"""
Exportación conservando:
- el formato visual del mayor original en la hoja Conciliación
- filas de separación y suma movimientos
- columnas extra de conciliación
- el signo y la columna original del mayor
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    'header_bg': '14444E',
    'header_font': 'FFFFFF',
    'matched_bg': 'D9F0EC',
    'one_n_bg': 'D4E8F7',
    'greedy_bg': 'FAEADF',
    'pending_bg': 'F8D7D3',
    'suggested_bg': 'FFF3CD',
    'account_bg': 'D9D9D9',
    'border': 'C8CDD3',
    'text_dark': '14444E',
}
thin_border = Border(
    left=Side(style='thin', color=C['border']),
    right=Side(style='thin', color=C['border']),
    top=Side(style='thin', color=C['border']),
    bottom=Side(style='thin', color=C['border'])
)

hdr_font = Font(name='Calibri', bold=True, color=C['header_font'], size=11)
hdr_fill = PatternFill('solid', fgColor=C['header_bg'])
matched_fill = PatternFill('solid', fgColor=C['matched_bg'])
one_n_fill = PatternFill('solid', fgColor=C['one_n_bg'])
greedy_fill = PatternFill('solid', fgColor=C['greedy_bg'])
unmatched_fill = PatternFill('solid', fgColor=C['pending_bg'])
suggested_fill = PatternFill('solid', fgColor=C['suggested_bg'])
account_fill = PatternFill('solid', fgColor=C['account_bg'])
acc_font = Font(name='Calibri', bold=True, size=11, color='000000')
money_fmt = '#,##0.00;[Red]-#,##0.00'
date_fmt = 'dd/mm/yyyy'

HEADERS = ['Cuenta', 'Nombre', 'Fecha', 'Asiento', 'Contrapartida',
           'Descripción', 'Debe', 'Haber', 'Estado', 'ID Match',
           'Tipo Match', 'Confianza', 'Tipo incidencia', 'Relacionado']
COL_WIDTHS = [14, 28, 14, 12, 14, 40, 14, 14, 13, 10, 13, 10, 18, 28]
DETAIL_HEADERS = ['Cuenta', 'Nombre', 'Mov. Debe', 'Mov. Haber', 'Total Debe',
                  'Total Haber', 'Saldo', 'Conciliados', 'Pendientes', '% Conciliación', 'Incidencias']


def _write_headers(ws):
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _match_fill(mov, has_both, match_lookup):
    if mov.match_id:
        mi = match_lookup.get((mov.side, mov.row_idx))
        mt = (mi.match_type if mi else '') or ''
        if 'greedy' in mt:
            return greedy_fill
        if mt in ('1:N', 'N:1') or mov.match_type in ('1:N_parent', '1:N_child'):
            return one_n_fill
        return matched_fill
    if getattr(mov, 'incidence_type', None) or getattr(mov, 'related_info', None):
        return suggested_fill
    return unmatched_fill


def _write_visual_row(ws, row_num, values, fill=None):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.border = thin_border
        if ci == 3 and val is not None and hasattr(val, 'year'):
            cell.number_format = date_fmt
        if ci in (7, 8):
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal='right')
        if fill:
            cell.fill = fill
    return row_num + 1


def _write_movements(ws, accounts, match_lookup, only_pending=False):
    current_row = 2
    for acc in accounts:
        has_both = (any(m.side == 'debe' for m in acc.movements) and
                    any(m.side == 'haber' for m in acc.movements))
        is_financial = str(acc.codigo).startswith('570') or str(acc.codigo).startswith('572')
        if not acc.movements:
            continue

        movs = list(acc.movements)
        if only_pending:
            if not has_both or is_financial:
                continue
            movs = [m for m in acc.movements if m.match_id is None]
            if not movs:
                continue

        # Cabecera cuenta
        ws.cell(row=current_row, column=1, value=acc.codigo)
        ws.cell(row=current_row, column=2, value=acc.nombre)
        for ci in range(1, len(HEADERS) + 1):
            c = ws.cell(row=current_row, column=ci)
            c.fill = account_fill
            c.font = acc_font
            c.border = thin_border
        current_row += 1

        # Fila en blanco
        current_row += 1

        sum_debe_visual = 0.0
        sum_haber_visual = 0.0

        for mov in movs:
            fecha_val = mov.fecha
            if fecha_val is not None and pd.notna(fecha_val):
                try:
                    fecha_val = pd.to_datetime(fecha_val)
                except Exception:
                    pass
            else:
                fecha_val = None

            # Mantener EXACTAMENTE la columna y el signo originales del mayor
            debe_val = None
            haber_val = None
            if getattr(mov, 'original_side', mov.side) == 'debe':
                debe_val = getattr(mov, 'original_importe', mov.importe)
                sum_debe_visual += float(debe_val or 0)
            else:
                haber_val = getattr(mov, 'original_importe', mov.importe)
                sum_haber_visual += float(haber_val or 0)

            mi = match_lookup.get((mov.side, mov.row_idx))
            tipo_match = mi.match_type if mi else ''
            confianza = mi.confidence if mi else ''

            row_data = [
                '', '', fecha_val, mov.asiento, mov.contrapartida, mov.descripcion,
                debe_val, haber_val,
                'Conciliado' if mov.match_id else 'Pendiente',
                mov.match_id or '', tipo_match or '', confianza or '',
                getattr(mov, 'incidence_type', '') or '',
                getattr(mov, 'related_info', '') or '',
            ]
            fill = _match_fill(mov, has_both, match_lookup)
            current_row = _write_visual_row(ws, current_row, row_data, fill=fill)

        sum_values = [
            '', 'Suma Movimientos ...', None, None, None, None,
            round(sum_debe_visual, 2),
            round(sum_haber_visual, 2),
            '', '', '', '', '', ''
        ]
        current_row = _write_visual_row(ws, current_row, sum_values, fill=account_fill)

        current_row += 1


def export_reconciliation(accounts, all_matches, summary, output_path=None):
    wb = Workbook()

    match_lookup = {}
    for m in all_matches:
        for di in m.debe_indices:
            match_lookup[('debe', di)] = m
        for hi in m.haber_indices:
            match_lookup[('haber', hi)] = m

    ws1 = wb.active
    ws1.title = "Conciliación"
    _write_headers(ws1)
    _write_movements(ws1, accounts, match_lookup, only_pending=False)

    ws2 = wb.create_sheet("Pendientes")
    _write_headers(ws2)
    _write_movements(ws2, accounts, match_lookup, only_pending=True)

    ws3 = wb.create_sheet("Informe")
    ws3.merge_cells('A1:F1')
    title_cell = ws3.cell(row=1, column=1, value='INFORME DE CONCILIACIÓN CONTABLE')
    title_cell.font = Font(name='Calibri', bold=True, size=16, color=C['text_dark'])
    title_cell.alignment = Alignment(horizontal='center')

    data = [
        ('', ''), ('RESUMEN GENERAL', ''),
        ('Total de cuentas detectadas', summary.get('total_accounts', 0)),
        ('Cuentas analizadas (debe y haber)', summary.get('analyzed_accounts', summary.get('accounts_with_both_sides', 0))),
        ('Cuentas excluidas', summary.get('excluded_accounts', 0)),
        ('', ''), ('MOVIMIENTOS', ''),
        ('Movimientos analizados', summary.get('analyzed_movements', summary.get('total_movements', 0))),
        ('Movimientos conciliados', summary.get('total_matched', 0)),
        ('Movimientos pendientes', summary.get('total_unmatched', 0)),
        ('Tasa de conciliación', f"{summary.get('match_rate', 0)}%"),
        ('', ''), ('IMPORTES', ''),
        ('Total Debe', summary.get('total_debe', 0)),
        ('Total Haber', summary.get('total_haber', 0)),
        ('Saldo', summary.get('balance', 0)),
        ('', ''), ('ESTADO DE CUENTAS', ''),
        ('Totalmente conciliadas (≥95%)', summary.get('fully_reconciled', 0)),
        ('Parcialmente conciliadas', summary.get('partially_reconciled', 0)),
        ('Sin conciliar', summary.get('not_reconciled', 0)),
    ]
    sec_font = Font(name='Calibri', bold=True, size=12, color=C['text_dark'])
    val_font = Font(name='Calibri', size=11)
    for r, (label, value) in enumerate(data, 3):
        cl = ws3.cell(row=r, column=1, value=label)
        cv = ws3.cell(row=r, column=2, value=value)
        if value == '' and label:
            cl.font = sec_font
        else:
            cl.font = val_font
            cv.font = Font(name='Calibri', size=11, bold=True)
            if isinstance(value, float):
                cv.number_format = money_fmt
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 20

    ws4 = wb.create_sheet("Detalle Cuentas")
    for ci, h in enumerate(DETAIL_HEADERS, 1):
        cell = ws4.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws4.freeze_panes = 'A2'

    rn = 2
    for acc in accounts:
        if not acc.movements:
            continue
        is_financial = str(acc.codigo).startswith('570') or str(acc.codigo).startswith('572')
        has_both = any(m.side == 'debe' for m in acc.movements) and any(m.side == 'haber' for m in acc.movements)
        analyzed = has_both and not is_financial
        matched = sum(1 for m in acc.movements if m.match_id is not None)
        vals = [
            acc.codigo, acc.nombre,
            sum(1 for m in acc.movements if m.side == 'debe'),
            sum(1 for m in acc.movements if m.side == 'haber'),
            acc.debe_total, acc.haber_total, acc.saldo,
            matched if analyzed else '',
            (len(acc.movements) - matched) if analyzed else '',
            round((matched / len(acc.movements)) * 100, 1) if analyzed and acc.movements else '',
            sum(1 for m in acc.movements if getattr(m, 'incidence_type', None))
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws4.cell(row=rn, column=ci, value=val)
            cell.border = thin_border
            if ci in (5, 6, 7):
                cell.number_format = money_fmt
        rn += 1

    for i, w in enumerate([14, 32, 12, 12, 14, 14, 14, 12, 12, 15, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    out = output_path if output_path else io.BytesIO()
    wb.save(out)
    if output_path:
        return output_path
    out.seek(0)
    return out
